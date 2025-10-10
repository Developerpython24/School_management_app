from io import BytesIO
import jdatetime
from flask import send_file
from openpyxl import Workbook  # فیکس: openpyxl مستقیم بدون pandas
from models import Score, SkillScore, Student, Class

def generate_excel_report(student, scores, skills, period):
    output = BytesIO()
    wb = Workbook()
    
    # Sheet نمرات درسی
    ws_scores = wb.active
    ws_scores.title = 'نمرات درسی'
    ws_scores.append(['تاریخ', 'روز هفته', 'درس', 'نمره'])  # header
    for score in scores:
        jdate = jdatetime.date.fromgregorian(date=score.date).strftime('%Y/%m/%d')
        ws_scores.append([jdate, score.weekday, score.subject_ref.name, score.score])
    
    # Sheet نمرات مهارتی (اگر وجود داره)
    if skills:
        ws_skills = wb.create_sheet('نمرات مهارتی')
        ws_skills.append(['تاریخ', 'مهارت', 'نمره'])
        for skill in skills:
            jdate = jdatetime.date.fromgregorian(date=skill.date).strftime('%Y/%m/%d')
            ws_skills.append([jdate, skill.skill_name, skill.score])
    
    wb.save(output)
    output.seek(0)
    filename = f'report_{student.first_name}_{student.last_name}_{period}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def generate_class_excel_report(class_obj, students, period, include_skills):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f'کلاس {class_obj.name}'
    ws.append(['کد دانش آموزی', 'نام', 'نام خانوادگی', 'میانگین نمرات'])  # header
    
    for student in students:
        scores = Score.query.filter_by(student_id=student.id).all()
        avg_score = sum(s.score for s in scores) / len(scores) if scores else 0
        row = [student.student_id, student.first_name, student.last_name, round(avg_score, 2)]
        ws.append(row)
        
        if include_skills:
            skills = SkillScore.query.filter_by(student_id=student.id).all()
            avg_skill = sum(s.score for s in skills) / len(skills) if skills else 0
            ws.append(['', '', '', f'میانگین مهارت: {round(avg_skill, 2)}'])  # row اضافی برای مهارت
    
    wb.save(output)
    output.seek(0)
    filename = f'class_report_{class_obj.name}_{period}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
